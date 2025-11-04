# -*- coding: utf-8 -*-
"""
@文件: report_generator.py
@說明: 合規報告生成器
@時間: 2025-01-09
@作者: LiDong
"""

import os
import json
import csv
import uuid
from datetime import datetime, date
from typing import Dict, List, Any, Optional, Tuple
from io import StringIO, BytesIO

import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from common.common_tools import CommonTools
from models.audit_model import OperAuditLogModel, OperSecurityEventModel
from loggers import logger


class ComplianceReportGenerator:
    """合规报告生成器"""
    
    def __init__(self):
        self.oper_audit_log = OperAuditLogModel()
        self.oper_security_event = OperSecurityEventModel()
        self.report_base_path = "reports"
        
        # 确保报告目录存在
        os.makedirs(self.report_base_path, exist_ok=True)
    
    def generate_audit_compliance_report(self, report_type: str, period_start: date, 
                                       period_end: date, format_type: str = "pdf") -> Tuple[str, Dict[str, Any]]:
        """生成审计合规报告"""
        try:
            # 收集报告数据
            report_data = self._collect_audit_data(period_start, period_end)
            
            # 根据报告类型处理数据
            if report_type == "access_control":
                processed_data = self._process_access_control_data(report_data)
            elif report_type == "data_access":
                processed_data = self._process_data_access_data(report_data)
            elif report_type == "security_incident":
                processed_data = self._process_security_incident_data(report_data)
            elif report_type == "compliance_summary":
                processed_data = self._process_compliance_summary_data(report_data)
            else:
                processed_data = self._process_general_audit_data(report_data)
            
            # 生成报告文件
            if format_type.lower() == "pdf":
                file_path = self._generate_pdf_report(processed_data, report_type, period_start, period_end)
            elif format_type.lower() == "excel":
                file_path = self._generate_excel_report(processed_data, report_type, period_start, period_end)
            elif format_type.lower() == "csv":
                file_path = self._generate_csv_report(processed_data, report_type, period_start, period_end)
            else:
                file_path = self._generate_json_report(processed_data, report_type, period_start, period_end)
            
            return file_path, processed_data
            
        except Exception as e:
            logger.error(f"生成合規報告失敗: {str(e)}")
            raise Exception(f"生成合規報告失敗: {str(e)}")
    
    def _collect_audit_data(self, period_start: date, period_end: date) -> Dict[str, Any]:
        """收集审计数据"""
        try:
            filters = {
                'start_time': period_start.strftime('%Y-%m-%d'),
                'end_time': period_end.strftime('%Y-%m-%d')
            }
            
            # 获取审计日志
            audit_logs = self.oper_audit_log.get_audit_logs(filters, page=1, per_page=10000)
            
            # 获取统计数据
            statistics = self.oper_audit_log.get_audit_statistics(filters)
            
            # 获取安全事件
            security_events = self.oper_security_event.get_security_events(filters, page=1, per_page=1000)
            
            return {
                'audit_logs': audit_logs,
                'statistics': statistics,
                'security_events': security_events,
                'period': {
                    'start': period_start.strftime('%Y-%m-%d'),
                    'end': period_end.strftime('%Y-%m-%d')
                }
            }
            
        except Exception as e:
            logger.error(f"收集審計數據失敗: {str(e)}")
            return {}
    
    def _process_access_control_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """处理访问控制数据"""
        audit_logs = data.get('audit_logs', [])
        
        # 按用户统计访问次数
        user_access = {}
        failed_access = []
        privileged_operations = []
        
        for log in audit_logs:
            user_id = log.get('user_id', 'anonymous')
            action = log.get('action', '')
            result = log.get('result', '')
            
            # 用户访问统计
            if user_id not in user_access:
                user_access[user_id] = {'total': 0, 'success': 0, 'failure': 0}
            user_access[user_id]['total'] += 1
            if result == 'success':
                user_access[user_id]['success'] += 1
            else:
                user_access[user_id]['failure'] += 1
            
            # 失败访问记录
            if result in ['failure', 'error']:
                failed_access.append({
                    'user_id': user_id,
                    'action': action,
                    'timestamp': log.get('timestamp'),
                    'ip_address': log.get('ip_address'),
                    'error_message': log.get('error_message')
                })
            
            # 特权操作记录
            if any(keyword in action.lower() for keyword in ['admin', 'delete', 'modify', 'config']):
                privileged_operations.append({
                    'user_id': user_id,
                    'action': action,
                    'resource_type': log.get('resource_type'),
                    'timestamp': log.get('timestamp'),
                    'result': result
                })
        
        return {
            'report_type': 'access_control',
            'summary': {
                'total_users': len(user_access),
                'total_access_attempts': sum(ua['total'] for ua in user_access.values()),
                'failed_attempts': len(failed_access),
                'privileged_operations': len(privileged_operations)
            },
            'user_access_stats': user_access,
            'failed_access': failed_access[:100],  # 限制显示数量
            'privileged_operations': privileged_operations[:100],
            'period': data.get('period', {})
        }
    
    def _process_data_access_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """处理数据访问数据"""
        audit_logs = data.get('audit_logs', [])
        
        # 按资源类型统计
        resource_access = {}
        sensitive_operations = []
        data_modifications = []
        
        for log in audit_logs:
            resource_type = log.get('resource_type', 'unknown')
            action = log.get('action', '')
            
            # 资源访问统计
            if resource_type not in resource_access:
                resource_access[resource_type] = {'read': 0, 'write': 0, 'delete': 0, 'other': 0}
            
            if 'read' in action.lower() or 'get' in action.lower():
                resource_access[resource_type]['read'] += 1
            elif 'write' in action.lower() or 'create' in action.lower() or 'update' in action.lower():
                resource_access[resource_type]['write'] += 1
            elif 'delete' in action.lower():
                resource_access[resource_type]['delete'] += 1
            else:
                resource_access[resource_type]['other'] += 1
            
            # 敏感操作
            if log.get('risk_level') in ['high', 'critical']:
                sensitive_operations.append({
                    'user_id': log.get('user_id'),
                    'action': action,
                    'resource_type': resource_type,
                    'resource_id': log.get('resource_id'),
                    'risk_level': log.get('risk_level'),
                    'timestamp': log.get('timestamp')
                })
            
            # 数据修改记录
            if log.get('old_values') or log.get('new_values'):
                data_modifications.append({
                    'user_id': log.get('user_id'),
                    'action': action,
                    'resource_type': resource_type,
                    'resource_id': log.get('resource_id'),
                    'has_old_values': bool(log.get('old_values')),
                    'has_new_values': bool(log.get('new_values')),
                    'timestamp': log.get('timestamp')
                })
        
        return {
            'report_type': 'data_access',
            'summary': {
                'total_resources': len(resource_access),
                'sensitive_operations': len(sensitive_operations),
                'data_modifications': len(data_modifications)
            },
            'resource_access_stats': resource_access,
            'sensitive_operations': sensitive_operations[:100],
            'data_modifications': data_modifications[:100],
            'period': data.get('period', {})
        }
    
    def _process_security_incident_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """处理安全事件数据"""
        security_events = data.get('security_events', [])
        
        # 按严重程度统计
        severity_stats = {'low': 0, 'medium': 0, 'high': 0, 'critical': 0}
        status_stats = {'open': 0, 'investigating': 0, 'resolved': 0, 'false_positive': 0}
        event_types = {}
        
        for event in security_events:
            severity = event.get('severity', 'low')
            status = event.get('status', 'open')
            event_type = event.get('event_type', 'unknown')
            
            severity_stats[severity] = severity_stats.get(severity, 0) + 1
            status_stats[status] = status_stats.get(status, 0) + 1
            event_types[event_type] = event_types.get(event_type, 0) + 1
        
        return {
            'report_type': 'security_incident',
            'summary': {
                'total_events': len(security_events),
                'critical_events': severity_stats.get('critical', 0),
                'high_severity_events': severity_stats.get('high', 0),
                'open_events': status_stats.get('open', 0),
                'resolved_events': status_stats.get('resolved', 0)
            },
            'severity_stats': severity_stats,
            'status_stats': status_stats,
            'event_type_stats': event_types,
            'events': security_events[:50],  # 限制显示数量
            'period': data.get('period', {})
        }
    
    def _process_compliance_summary_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """处理合规总结数据"""
        statistics = data.get('statistics', {})
        audit_logs = data.get('audit_logs', [])
        security_events = data.get('security_events', [])
        
        # 计算合规指标
        total_operations = len(audit_logs)
        failed_operations = len([log for log in audit_logs if log.get('result') in ['failure', 'error']])
        high_risk_operations = len([log for log in audit_logs if log.get('risk_level') in ['high', 'critical']])
        
        security_score = self._calculate_security_score(audit_logs, security_events)
        compliance_score = self._calculate_compliance_score(audit_logs, security_events)
        
        return {
            'report_type': 'compliance_summary',
            'summary': {
                'total_operations': total_operations,
                'failed_operations': failed_operations,
                'failure_rate': (failed_operations / total_operations * 100) if total_operations > 0 else 0,
                'high_risk_operations': high_risk_operations,
                'security_score': security_score,
                'compliance_score': compliance_score
            },
            'statistics': statistics,
            'recommendations': self._generate_recommendations(audit_logs, security_events),
            'period': data.get('period', {})
        }
    
    def _process_general_audit_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """处理通用审计数据"""
        return {
            'report_type': 'general_audit',
            'summary': data.get('statistics', {}),
            'audit_logs': data.get('audit_logs', [])[:100],
            'security_events': data.get('security_events', [])[:50],
            'period': data.get('period', {})
        }
    
    def _calculate_security_score(self, audit_logs: List[Dict], security_events: List[Dict]) -> float:
        """计算安全评分"""
        try:
            total_score = 100.0
            
            # 根据失败率扣分
            total_ops = len(audit_logs)
            if total_ops > 0:
                failed_ops = len([log for log in audit_logs if log.get('result') in ['failure', 'error']])
                failure_rate = failed_ops / total_ops
                total_score -= failure_rate * 30  # 最多扣30分
            
            # 根据高风险操作扣分
            high_risk_ops = len([log for log in audit_logs if log.get('risk_level') in ['high', 'critical']])
            if high_risk_ops > 0:
                risk_ratio = min(high_risk_ops / max(total_ops, 1), 0.5)
                total_score -= risk_ratio * 20  # 最多扣20分
            
            # 根据安全事件扣分
            critical_events = len([event for event in security_events if event.get('severity') == 'critical'])
            high_events = len([event for event in security_events if event.get('severity') == 'high'])
            total_score -= critical_events * 10  # 每个严重事件扣10分
            total_score -= high_events * 5     # 每个高级事件扣5分
            
            return max(total_score, 0.0)
            
        except Exception as e:
            logger.error(f"計算安全評分失敗: {str(e)}")
            return 50.0
    
    def _calculate_compliance_score(self, audit_logs: List[Dict], security_events: List[Dict]) -> float:
        """计算合规评分"""
        try:
            total_score = 100.0
            
            # 审计日志完整性检查
            incomplete_logs = len([log for log in audit_logs if not log.get('user_id') or not log.get('action')])
            if len(audit_logs) > 0:
                incomplete_ratio = incomplete_logs / len(audit_logs)
                total_score -= incomplete_ratio * 25  # 最多扣25分
            
            # 响应时间检查
            unresolved_events = len([event for event in security_events if event.get('status') == 'open'])
            if unresolved_events > 0:
                total_score -= min(unresolved_events * 5, 25)  # 最多扣25分
            
            # 权限使用检查
            privileged_ops = len([log for log in audit_logs if 'admin' in log.get('action', '').lower()])
            if len(audit_logs) > 0:
                privileged_ratio = privileged_ops / len(audit_logs)
                if privileged_ratio > 0.1:  # 超过10%为异常
                    total_score -= (privileged_ratio - 0.1) * 100
            
            return max(total_score, 0.0)
            
        except Exception as e:
            logger.error(f"計算合規評分失敗: {str(e)}")
            return 50.0
    
    def _generate_recommendations(self, audit_logs: List[Dict], security_events: List[Dict]) -> List[str]:
        """生成建议"""
        recommendations = []
        
        # 基于审计日志的建议
        if audit_logs:
            failed_rate = len([log for log in audit_logs if log.get('result') in ['failure', 'error']]) / len(audit_logs)
            if failed_rate > 0.1:
                recommendations.append("失敗操作比例過高，建議檢查系統配置和用戶權限")
            
            high_risk_rate = len([log for log in audit_logs if log.get('risk_level') in ['high', 'critical']]) / len(audit_logs)
            if high_risk_rate > 0.05:
                recommendations.append("高風險操作比例較高，建議加強審核流程")
        
        # 基于安全事件的建议
        if security_events:
            open_events = [event for event in security_events if event.get('status') == 'open']
            if len(open_events) > 5:
                recommendations.append("存在多個未處理的安全事件，建議優先處理")
            
            critical_events = [event for event in security_events if event.get('severity') == 'critical']
            if critical_events:
                recommendations.append("存在嚴重安全事件，建議立即處理並加強監控")
        
        if not recommendations:
            recommendations.append("系統運行正常，建議繼續保持現有的安全管理措施")
        
        return recommendations
    
    def _generate_pdf_report(self, data: Dict[str, Any], report_type: str, 
                           period_start: date, period_end: date) -> str:
        """生成PDF报告"""
        try:
            # 生成文件名
            filename = f"{report_type}_report_{period_start.strftime('%Y%m%d')}_{period_end.strftime('%Y%m%d')}_{uuid.uuid4().hex[:8]}.pdf"
            file_path = os.path.join(self.report_base_path, filename)
            
            # 创建PDF文档
            doc = SimpleDocTemplate(file_path, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # 标题
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # 居中对齐
            )
            title = f"合規報告 - {report_type.upper()}"
            story.append(Paragraph(title, title_style))
            
            # 报告期间
            period_text = f"報告期間: {period_start.strftime('%Y-%m-%d')} 至 {period_end.strftime('%Y-%m-%d')}"
            story.append(Paragraph(period_text, styles['Normal']))
            story.append(Spacer(1, 20))
            
            # 摘要信息
            if 'summary' in data:
                story.append(Paragraph("執行摘要", styles['Heading2']))
                summary_data = []
                for key, value in data['summary'].items():
                    summary_data.append([key.replace('_', ' ').title(), str(value)])
                
                summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
                summary_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(summary_table)
                story.append(Spacer(1, 20))
            
            # 建议（如果有）
            if 'recommendations' in data and data['recommendations']:
                story.append(Paragraph("建議事項", styles['Heading2']))
                for i, rec in enumerate(data['recommendations'], 1):
                    story.append(Paragraph(f"{i}. {rec}", styles['Normal']))
                story.append(Spacer(1, 20))
            
            # 生成时间
            generated_time = f"報告生成時間: {CommonTools.get_now()}"
            story.append(Paragraph(generated_time, styles['Normal']))
            
            # 构建PDF
            doc.build(story)
            
            return file_path
            
        except Exception as e:
            logger.error(f"生成PDF報告失敗: {str(e)}")
            raise Exception(f"生成PDF報告失敗: {str(e)}")
    
    def _generate_excel_report(self, data: Dict[str, Any], report_type: str,
                              period_start: date, period_end: date) -> str:
        """生成Excel报告"""
        try:
            # 生成文件名
            filename = f"{report_type}_report_{period_start.strftime('%Y%m%d')}_{period_end.strftime('%Y%m%d')}_{uuid.uuid4().hex[:8]}.xlsx"
            file_path = os.path.join(self.report_base_path, filename)
            
            # 创建工作簿
            wb = Workbook()
            
            # 摘要工作表
            ws_summary = wb.active
            ws_summary.title = "摘要"
            
            # 设置标题
            ws_summary['A1'] = f"合規報告 - {report_type.upper()}"
            ws_summary['A1'].font = Font(size=16, bold=True)
            ws_summary['A2'] = f"報告期間: {period_start.strftime('%Y-%m-%d')} 至 {period_end.strftime('%Y-%m-%d')}"
            
            # 摘要数据
            if 'summary' in data:
                row = 4
                ws_summary['A3'] = "執行摘要"
                ws_summary['A3'].font = Font(size=14, bold=True)
                
                for key, value in data['summary'].items():
                    ws_summary[f'A{row}'] = key.replace('_', ' ').title()
                    ws_summary[f'B{row}'] = str(value)
                    row += 1
            
            # 建议工作表
            if 'recommendations' in data and data['recommendations']:
                ws_rec = wb.create_sheet("建議事項")
                ws_rec['A1'] = "建議事項"
                ws_rec['A1'].font = Font(size=14, bold=True)
                
                for i, rec in enumerate(data['recommendations'], 2):
                    ws_rec[f'A{i}'] = f"{i-1}. {rec}"
            
            # 保存文件
            wb.save(file_path)
            
            return file_path
            
        except Exception as e:
            logger.error(f"生成Excel報告失敗: {str(e)}")
            raise Exception(f"生成Excel報告失敗: {str(e)}")
    
    def _generate_csv_report(self, data: Dict[str, Any], report_type: str,
                           period_start: date, period_end: date) -> str:
        """生成CSV报告"""
        try:
            # 生成文件名
            filename = f"{report_type}_report_{period_start.strftime('%Y%m%d')}_{period_end.strftime('%Y%m%d')}_{uuid.uuid4().hex[:8]}.csv"
            file_path = os.path.join(self.report_base_path, filename)
            
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                
                # 写入标题信息
                writer.writerow([f"合規報告 - {report_type.upper()}"])
                writer.writerow([f"報告期間: {period_start.strftime('%Y-%m-%d')} 至 {period_end.strftime('%Y-%m-%d')}"])
                writer.writerow([])  # 空行
                
                # 写入摘要
                if 'summary' in data:
                    writer.writerow(["執行摘要"])
                    writer.writerow(["指標", "數值"])
                    for key, value in data['summary'].items():
                        writer.writerow([key.replace('_', ' ').title(), str(value)])
                    writer.writerow([])  # 空行
                
                # 写入建议
                if 'recommendations' in data and data['recommendations']:
                    writer.writerow(["建議事項"])
                    for i, rec in enumerate(data['recommendations'], 1):
                        writer.writerow([f"{i}. {rec}"])
                
                # 写入生成时间
                writer.writerow([])
                writer.writerow([f"報告生成時間: {CommonTools.get_now()}"])
            
            return file_path
            
        except Exception as e:
            logger.error(f"生成CSV報告失敗: {str(e)}")
            raise Exception(f"生成CSV報告失敗: {str(e)}")
    
    def _generate_json_report(self, data: Dict[str, Any], report_type: str,
                            period_start: date, period_end: date) -> str:
        """生成JSON报告"""
        try:
            # 生成文件名
            filename = f"{report_type}_report_{period_start.strftime('%Y%m%d')}_{period_end.strftime('%Y%m%d')}_{uuid.uuid4().hex[:8]}.json"
            file_path = os.path.join(self.report_base_path, filename)
            
            # 添加元数据
            report_data = {
                'title': f"合規報告 - {report_type.upper()}",
                'report_type': report_type,
                'period': {
                    'start': period_start.strftime('%Y-%m-%d'),
                    'end': period_end.strftime('%Y-%m-%d')
                },
                'generated_at': CommonTools.get_now(),
                'data': data
            }
            
            with open(file_path, 'w', encoding='utf-8') as jsonfile:
                json.dump(report_data, jsonfile, ensure_ascii=False, indent=2)
            
            return file_path
            
        except Exception as e:
            logger.error(f"生成JSON報告失敗: {str(e)}")
            raise Exception(f"生成JSON報告失敗: {str(e)}")


# 全局报告生成器实例
report_generator = ComplianceReportGenerator()